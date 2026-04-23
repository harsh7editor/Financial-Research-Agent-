"""
Report Export Tool (Phase 3.3).

Generates PDF and Excel reports from analysis data using reportlab
and openpyxl. Supports three templates: executive summary, deep dive,
and portfolio review.

Usage::

    from src.tools.report_export import generate_pdf_report, generate_excel_report
    pdf_path = generate_pdf_report("AAPL", analyses)
    xlsx_path = generate_excel_report("AAPL", analyses)
"""

from __future__ import annotations

import io
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from src.utils.logger import get_logger

logger = get_logger(__name__)


def _get_output_dir() -> Path:
    """Get reports output directory."""
    from src.config import settings
    path = settings.data_dir / "reports"
    path.mkdir(parents=True, exist_ok=True)
    return path


# ── PDF Generation ───────────────────────────────────────────


def generate_pdf_report(
    symbol: str,
    analyses: Dict[str, Any],
    template: str = "executive_summary",
) -> Dict[str, Any]:
    """
    Generate a PDF research report.

    Args:
        symbol: Stock ticker.
        analyses: Dict with analysis results (technical, fundamental, etc.).
        template: "executive_summary", "deep_dive", or "portfolio_review".

    Returns:
        Dict with file_path, file_size, and metadata.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, mm
        from reportlab.lib.colors import HexColor
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, HRFlowable,
        )
        from reportlab.lib import colors

        output_dir = _get_output_dir()
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"{symbol}_{template}_{timestamp}.pdf"
        filepath = output_dir / filename

        doc = SimpleDocTemplate(
            str(filepath),
            pagesize=A4,
            rightMargin=20 * mm,
            leftMargin=20 * mm,
            topMargin=20 * mm,
            bottomMargin=20 * mm,
        )

        styles = getSampleStyleSheet()
        # Custom styles
        styles.add(ParagraphStyle(
            "ReportTitle", parent=styles["Title"],
            fontSize=22, spaceAfter=12,
            textColor=HexColor("#1a1a2e"),
        ))
        styles.add(ParagraphStyle(
            "SectionHeader", parent=styles["Heading2"],
            fontSize=14, spaceAfter=8, spaceBefore=16,
            textColor=HexColor("#16213e"),
        ))
        styles.add(ParagraphStyle(
            "SubHeader", parent=styles["Heading3"],
            fontSize=11, spaceAfter=6,
            textColor=HexColor("#0f3460"),
        ))
        styles.add(ParagraphStyle(
            "BodyText2", parent=styles["BodyText"],
            fontSize=10, spaceAfter=6,
            leading=14,
        ))
        styles.add(ParagraphStyle(
            "Disclaimer", parent=styles["Normal"],
            fontSize=8, textColor=HexColor("#888888"),
            spaceAfter=4,
        ))

        elements = []

        # ── Title Page ──
        elements.append(Spacer(1, 40 * mm))
        elements.append(Paragraph(f"Investment Research Report", styles["ReportTitle"]))
        elements.append(Paragraph(f"<b>{symbol}</b>", styles["ReportTitle"]))
        elements.append(Spacer(1, 10 * mm))
        elements.append(HRFlowable(width="80%", thickness=1, color=HexColor("#1a1a2e")))
        elements.append(Spacer(1, 5 * mm))

        company = analyses.get("company_info", {})
        elements.append(Paragraph(
            f"Company: {company.get('name', symbol)}<br/>"
            f"Sector: {company.get('sector', 'N/A')}<br/>"
            f"Date: {datetime.now(timezone.utc).strftime('%B %d, %Y')}<br/>"
            f"Analyst: AI Financial Research Agent",
            styles["BodyText2"],
        ))

        elements.append(PageBreak())

        # ── Executive Summary ──
        elements.append(Paragraph("Executive Summary", styles["SectionHeader"]))
        rec = analyses.get("recommendation", {})
        elements.append(Paragraph(
            f"<b>Recommendation:</b> {rec.get('action', 'HOLD')}<br/>"
            f"<b>Confidence:</b> {rec.get('confidence', 0.5) * 100:.0f}%<br/>"
            f"<b>Current Price:</b> ${analyses.get('current_price', 0):.2f}",
            styles["BodyText2"],
        ))
        elements.append(Spacer(1, 5 * mm))

        # ── Technical Analysis ──
        tech = analyses.get("technical", {})
        if tech:
            elements.append(Paragraph("Technical Analysis", styles["SectionHeader"]))
            tech_text = _format_analysis_section(tech)
            elements.append(Paragraph(tech_text, styles["BodyText2"]))

        # ── Fundamental Analysis ──
        fund = analyses.get("fundamental", {})
        if fund:
            elements.append(Paragraph("Fundamental Analysis", styles["SectionHeader"]))
            fund_text = _format_analysis_section(fund)
            elements.append(Paragraph(fund_text, styles["BodyText2"]))

        # ── Risk Assessment ──
        risk = analyses.get("risk", {})
        if risk:
            elements.append(Paragraph("Risk Assessment", styles["SectionHeader"]))
            risk_text = _format_analysis_section(risk)
            elements.append(Paragraph(risk_text, styles["BodyText2"]))

        # ── Sentiment ──
        sent = analyses.get("sentiment", {})
        if sent:
            elements.append(Paragraph("Sentiment Analysis", styles["SectionHeader"]))
            sent_text = _format_analysis_section(sent)
            elements.append(Paragraph(sent_text, styles["BodyText2"]))

        # ── DCF (if available) ──
        dcf = analyses.get("dcf", {})
        if dcf and "scenarios" in dcf:
            elements.append(Paragraph("DCF Valuation", styles["SectionHeader"]))
            scenarios = dcf["scenarios"]
            table_data = [["Scenario", "Intrinsic Value", "Upside/Downside"]]
            for name, data in scenarios.items():
                table_data.append([
                    name.title(),
                    f"${data.get('intrinsic_value', 0):.2f}",
                    f"{data.get('upside_pct', 0):+.1f}%",
                ])
            t = Table(table_data, colWidths=[80, 100, 100])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), HexColor("#16213e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]))
            elements.append(t)

        # ── Disclaimer ──
        elements.append(Spacer(1, 20 * mm))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=HexColor("#cccccc")))
        elements.append(Paragraph(
            "This report is for informational purposes only and does not constitute "
            "investment advice. Past performance does not guarantee future results. "
            "Generated by AI Financial Research Agent.",
            styles["Disclaimer"],
        ))

        doc.build(elements)

        file_size = filepath.stat().st_size

        return {
            "file_path": str(filepath),
            "filename": filename,
            "file_size_bytes": file_size,
            "file_size_kb": round(file_size / 1024, 1),
            "template": template,
            "symbol": symbol,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }

    except ImportError:
        return {"error": "reportlab not installed — pip install reportlab"}
    except Exception as e:
        logger.error(f"PDF generation failed for {symbol}: {e}")
        return {"error": str(e)}


# ── Excel Export ─────────────────────────────────────────────


def generate_excel_report(
    symbol: str,
    analyses: Dict[str, Any],
) -> Dict[str, Any]:
    """
    Generate an Excel report with multiple sheets.

    Sheets: Summary, Technical, Fundamental, Risk, Sentiment.

    Returns:
        Dict with file_path and metadata.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        output_dir = _get_output_dir()
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"{symbol}_report_{timestamp}.xlsx"
        filepath = output_dir / filename

        wb = Workbook()

        # Styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
        sub_font = Font(bold=True, size=10)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # ── Summary Sheet ──
        ws = wb.active
        ws.title = "Summary"
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30

        summary_data = [
            ("Symbol", symbol),
            ("Date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
            ("Current Price", f"${analyses.get('current_price', 0):.2f}"),
            ("Recommendation", analyses.get("recommendation", {}).get("action", "HOLD")),
            ("Confidence", f"{analyses.get('recommendation', {}).get('confidence', 0.5) * 100:.0f}%"),
        ]

        for r, (label, value) in enumerate(summary_data, 1):
            ws.cell(row=r, column=1, value=label).font = sub_font
            ws.cell(row=r, column=2, value=value)

        # ── Analysis sheets ──
        for sheet_name, data_key in [
            ("Technical", "technical"),
            ("Fundamental", "fundamental"),
            ("Risk", "risk"),
            ("Sentiment", "sentiment"),
        ]:
            ws = wb.create_sheet(sheet_name)
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 25

            data = analyses.get(data_key, {})
            _write_dict_to_sheet(ws, data, header_font, header_fill, thin_border)

        # ── DCF Sheet (if available) ──
        dcf = analyses.get("dcf", {})
        if dcf and "scenarios" in dcf:
            ws = wb.create_sheet("DCF Valuation")
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 15

            # Header
            for col, header in enumerate(["Scenario", "Intrinsic Value", "Upside %", "Growth Rate"], 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            for row, (name, sdata) in enumerate(dcf["scenarios"].items(), 2):
                ws.cell(row=row, column=1, value=name.title())
                ws.cell(row=row, column=2, value=f"${sdata.get('intrinsic_value', 0):.2f}")
                ws.cell(row=row, column=3, value=f"{sdata.get('upside_pct', 0):.1f}%")
                ws.cell(row=row, column=4, value=f"{sdata.get('growth_rate', 0):.1f}%")

        wb.save(str(filepath))

        file_size = filepath.stat().st_size

        return {
            "file_path": str(filepath),
            "filename": filename,
            "file_size_bytes": file_size,
            "file_size_kb": round(file_size / 1024, 1),
            "sheets": [ws.title for ws in wb.worksheets],
            "symbol": symbol,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        }

    except ImportError:
        return {"error": "openpyxl not installed — pip install openpyxl"}
    except Exception as e:
        logger.error(f"Excel generation failed for {symbol}: {e}")
        return {"error": str(e)}


# ── Helpers ──────────────────────────────────────────────────


def _format_analysis_section(data: Dict[str, Any], max_depth: int = 2) -> str:
    """Format analysis dict as readable HTML for reportlab Paragraphs."""
    lines = []
    for key, value in data.items():
        if key.startswith("_") or key in ("raw_result", "analyzed_at"):
            continue
        label = key.replace("_", " ").title()
        if isinstance(value, dict) and max_depth > 0:
            lines.append(f"<b>{label}:</b>")
            for k, v in value.items():
                if not str(k).startswith("_"):
                    lines.append(f"&nbsp;&nbsp;{k}: {_format_value(v)}")
        else:
            lines.append(f"<b>{label}:</b> {_format_value(value)}")
    return "<br/>".join(lines[:30])  # Limit to prevent oversized sections


def _format_value(val: Any) -> str:
    if isinstance(val, float):
        return f"{val:.2f}"
    if isinstance(val, list):
        return f"[{len(val)} items]"
    if isinstance(val, dict):
        return f"({len(val)} fields)"
    return str(val)[:100]


def _write_dict_to_sheet(ws, data: Dict, header_font, header_fill, border):
    """Write a flat dict to an Excel sheet."""
    row = 1
    for key, value in data.items():
        if key.startswith("_") or key in ("raw_result",):
            continue
        label = key.replace("_", " ").title()

        if isinstance(value, dict):
            # Section header
            cell = ws.cell(row=row, column=1, value=label)
            cell.font = header_font
            cell.fill = header_fill
            ws.cell(row=row, column=2).fill = header_fill
            row += 1
            for k, v in value.items():
                ws.cell(row=row, column=1, value=str(k).replace("_", " ").title())
                ws.cell(row=row, column=2, value=_format_value(v))
                row += 1
        else:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=_format_value(value))
            row += 1
